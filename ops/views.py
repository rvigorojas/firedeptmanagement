#coding=utf-8
from ops.forms import ServiceForm, AffectedForm, ServiceVehicleForm, ArrestForm, ArrestPaymentForm, ServiceImageForm
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.forms.formsets import formset_factory
from ops.models import ServiceVehicle, ServiceAffected, Service, ArrestPayment, ServiceImage
from personal.models import Firefighter
from common.models import BasePerson, TelephoneNumber, PersonTelephoneNumber
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.exceptions import ValidationError
from django.db import transaction, connection
from datetime import datetime, date, timedelta
from django.shortcuts import render
from django.db.models import Count
from django.db.models.functions import TruncMonth
from collections import OrderedDict as SortedDict
from django.http import HttpResponse
from itertools import groupby
from operator import itemgetter
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


import json
@login_required
def list_services(request):
    services_qs = Service.objects.all()
    paginator = Paginator(services_qs, 15)
    page = request.GET.get('page')
    try:
        services = paginator.page(page)
    except PageNotAnInteger:
        services = paginator.page(1)
    except EmptyPage:
        services = paginator.page(paginator.num_pages)

    return render(request, "list_services.html", {
        "services": services,
        "paginator": paginator,
        "service_type_choices": Service.SERVICE_TYPE_CHOICES,
    })


@login_required
def export_services_excel(request):
    """
    Exporta a Excel (.xlsx) los servicios (emergencias) registrados,
    opcionalmente filtrados por tipo de emergencia (?tipo=<codigo>).
    Agregado a partir de la retroalimentación de evaluación del proyecto.
    Ver docs/CORRECCIONES.md.
    """
    tipo = request.GET.get('tipo', '')
    services = Service.objects.all().order_by('-date', '-time')
    tipo_label = 'Todos los tipos'
    if tipo:
        services = services.filter(service_type=tipo)
        tipo_label = dict(Service.SERVICE_TYPE_CHOICES).get(tipo, tipo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"

    ws.append([u"Reporte de Servicios (Emergencias)"])
    ws.append([u"Tipo: %s" % tipo_label])
    ws.append([])

    headers = [u"ID", u"Fecha", u"Hora", u"Tipo", u"Descripción", u"Ubicación",
               u"Tiempo de Respuesta (min)", u"Duración (h)"]
    header_row = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for s in services:
        response_time = s.response_time()
        duration = s.duration()
        ws.append([
            s.id,
            s.date.strftime("%d/%m/%Y") if s.date else "",
            s.time.strftime("%H:%M") if s.time else "",
            s.get_service_type_display(),
            s.description,
            s.location or "",
            round(response_time.total_seconds() / 60.0, 1) if response_time else "",
            round(duration.total_seconds() / 3600.0, 2) if duration else "",
        ])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    filename = "servicios_%s.xlsx" % (tipo if tipo else "todos")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    wb.save(response)
    return response

@login_required
def view_service(request, service_id):
    service = Service.objects.get(id=service_id)
    try:
        allow_file_upload = request.user.firefighter in service.complete_crew()
    except:
        allow_file_upload = False
    data = {"service": service, 'allow_file_upload':allow_file_upload}
    
    if allow_file_upload:
        image_upload_form =  ServiceImageForm()
        data["image_upload_form"] = image_upload_form
    
    return render(request, "view_service.html", data)

@login_required
def service_upload_image(request, service_id):
    
    if request.method == 'POST':
        form = ServiceImageForm(request.POST, request.FILES)
        if form.is_valid():
            image = ServiceImage()
            image.service_id = service_id
            image.uploader = request.user.firefighter
            image.file = form.cleaned_data['image']
            image.save()
            messages.success(request, u'La foto se guardó exitosamente')
        else:
            messages.error(request, u'La foto no se guardó')
    return redirect('view_service', service_id=service_id)
    
    
@login_required
@transaction.atomic
def insert_service(request):
    params = {}
    AffectedFormSet = formset_factory(AffectedForm, extra=0)
    VehiclesFormSet = formset_factory(ServiceVehicleForm, extra=1)
    firefighter = request.user.firefighter
    service_form = ServiceForm()
    crew_dict = {}

    if request.method == 'POST':
        data = request.POST.copy()
        data['time'] = data['time'][0:-2]+":"+data['time'][-2:]
        data['end_time'] = data['end_time'][0:-2]+":"+data['end_time'][-2:]
        data['scene_arrival_time'] = data['scene_arrival_time'][0:-2]+":"+data['scene_arrival_time'][-2:]
        
        service_form = ServiceForm(data)
        affected_formset = AffectedFormSet(data, prefix='affected')
        vehicles_formset = VehiclesFormSet(data, prefix='vehicles')
        if service_form.is_valid() and affected_formset.is_valid() and vehicles_formset.is_valid():
            service = service_form.save()
            service.created_by = firefighter
            service.save()

            for vf in vehicles_formset.forms:
                if "lead" in vf.cleaned_data:
                    s_vehicle = ServiceVehicle()
                    s_vehicle.service = service
                    s_vehicle.lead = Firefighter.objects.get(id=vf.cleaned_data['lead'])
                    s_vehicle.vehicle = vf.cleaned_data['vehicle']
                    if vf.cleaned_data['driver']:
                        s_vehicle.driver = Firefighter.objects.get(id=vf.cleaned_data['driver'])

                    s_vehicle.save()

                    for crew_id in vf.cleaned_data['crew_ids'].split(","):
                        if crew_id != "":
                            s_vehicle.crew.add(Firefighter.objects.get(id=crew_id))

            for af in affected_formset.forms:
                if 'first_name' in af.cleaned_data:
                    if af.cleaned_data["id_document"]:
                        person, _ = BasePerson.objects.get_or_create(id_document=af.cleaned_data["id_document"])
                    else:
                        person = BasePerson()
                    person.first_name = af.cleaned_data['first_name']
                    person.first_name_2 = af.cleaned_data['first_name_2']
                    person.last_name = af.cleaned_data['last_name']
                    person.last_name_2 = af.cleaned_data['last_name_2']
                    person.gender = af.cleaned_data['gender']
                    if af.cleaned_data["primary_email"]:
                        person.primary_email = af.cleaned_data["primary_email"]

                    person.save()
                    if af.cleaned_data["phone_code"] and af.cleaned_data["phone_number"]:
                        telephone = TelephoneNumber(code=af.cleaned_data["phone_code"],
                                                    number=af.cleaned_data["phone_number"])
                        telephone.save()
                        PersonTelephoneNumber(person=person,type='O',
                                              telephone_number=telephone).save()

                    s_affected = ServiceAffected(person_affected=person,
                                                 notes=af.cleaned_data["notes"],
                                                 type=af.cleaned_data["type"])
                    s_affected.save()
                    service.affected.add(s_affected)
            messages.success(request, u'El servicio fue guardado exitosamente')
            return redirect(list_services)
        else:
            crew_ids_str = ""
            for k, v in data.items():
                if "crew_ids" in k and v!="":
                    crew_ids_str = crew_ids_str+","+v
            crew_ids = [x for x in crew_ids_str.split(",") if x!='']
            crew = Firefighter.objects.filter(id__in=crew_ids)
            for member in crew:
                crew_dict[member.id] = str(member)
    else:
        affected_formset = AffectedFormSet(prefix='affected')
        vehicles_formset = VehiclesFormSet(prefix='vehicles')
        
    params['form'] = service_form
    params['affected'] = affected_formset
    params['vehicles'] =  vehicles_formset
    params['media'] = service_form.media
    params['ff'] = firefighter
    params['crew_data'] = json.dumps(crew_dict)

    return render(request, "insert_service.html", params)


@login_required
@transaction.atomic
def insert_arrest(request):
    params = {}
    arrest_form = ArrestForm()
    firefighter = request.user.firefighter
    
    if request.method == 'POST':
        data = request.POST.copy()
        arrest_form = ArrestForm(data)
        if arrest_form.is_valid():
            arrest = arrest_form.save(commit=False)
            arrest.created_by = firefighter
            arrest.arrested = Firefighter.objects.get(id=arrest_form.cleaned_data['arrested'])
            arrest.save()
            messages.success(request, u'El arresto fue guardado exitosamente')
            return redirect(insert_arrest)

    params['arrest_form'] = arrest_form    
    return render(request, "insert_arrest.html", params)


@login_required
@transaction.atomic
def insert_arrest_payment(request):
    params = {}
    arrest_payment_form = ArrestPaymentForm()
    firefighter = request.user.firefighter
    
    if request.method == 'POST':
        data = request.POST.copy()
        data['start_time_time'] = data['start_time_time'][0:-2]+":"+data['start_time_time'][-2:]
        data['end_time_time'] = data['end_time_time'][0:-2]+":"+data['end_time_time'][-2:]
        
        arrest_payment_form = ArrestPaymentForm(data)
        if arrest_payment_form.is_valid():
            arrest_payment = ArrestPayment()
            arrest_payment.created_by = firefighter
            arrest_payment.start_time = datetime.combine(arrest_payment_form.cleaned_data['start_time_date'], arrest_payment_form.cleaned_data['start_time_time'])
            arrest_payment.end_time = datetime.combine(arrest_payment_form.cleaned_data['end_time_date'], arrest_payment_form.cleaned_data['end_time_time'])
            arrest_payment.payer = Firefighter.objects.get(id=arrest_payment_form.cleaned_data['payer'])
            try:
                arrest_payment.full_clean()
                arrest_payment.save()
                messages.success(request, u'El pago arresto fue guardado exitosamente')
                return redirect(insert_arrest_payment)
            except ValidationError as e:
                messages.error(request, e.message_dict["__all__"][0])
                

    params['arrest_payment_form'] = arrest_payment_form    
    return render(request, "insert_arrest_payment.html", params)

def statistics(request):
    data = stats_data(request)
    return render(request, 'statistics.html', data)

def plain_statistics(request):
    data = stats_data(request)
    return render(request, 'plain_statistics.html', data)



def stats_data(request):
    summary = Service.objects.filter(date__gt=date(year=2012, month=12, day=31)).annotate(month=TruncMonth('date')).values('month', 'service_type').annotate(Count('service_type')).order_by('-month', 'service_type')
    services_by_month = SortedDict()
    for k, g in  groupby(summary, itemgetter('month')):
        date_formatted = k.strftime("%Y-%m") if hasattr(k, 'strftime') else str(k).rsplit("-", 1)[0]
        total = 0
        for info in g:
            new_data= {"type":info['service_type'], 'count':info['service_type__count']}
            total+=info['service_type__count']
            if not date_formatted in services_by_month:
                services_by_month[date_formatted] = []
            services_by_month[date_formatted].append(new_data)
        services_by_month[date_formatted].append({"type":"Total:", 'count':total})
    data = {'services_by_month': services_by_month, 'type_legend': Service.SERVICE_TYPE_CHOICES, 'ga': settings.GA}
    return data



def month_statistics(request, year, month):
    services = Service.objects.filter(date__year=int(year)).filter(date__month=int(month)).values('service_type').annotate(Count('service_type')).order_by('service_type')
    data = {'info': [['Tipo', 'Cantidad']]+[[x['service_type'],x['service_type__count']]  for x in services]}
    return HttpResponse(json.dumps(data))


def month_statistics_detail(request, year, month):
    services = Service.objects.filter(date__year=int(year)).filter(date__month=int(month))
    
    response_times = list(filter(lambda x: x != None, (service.response_time() for service in services)))
    response_time = (sum(response_times, timedelta(0)).seconds/float(len(response_times)))/60.00 if len(response_times) else 0.0
    response_time_text =  ("%.1f" % response_time)+" minutos" if len(response_times) else "No Disponible" 
    
    durations = list(filter(lambda x: x != None, (service.duration() for service in services)))
    
    average_duration = (sum(durations, timedelta(0)).seconds/float(len(durations)))/3600.00 if len(durations) else 0.00
    average_duration_text =  ("%.2f" % average_duration)+" horas" if len(durations) else "No Disponible"
    
    time_in_service = (sum(durations, timedelta(0)).seconds)/3600.00
    time_in_service_text = ("%.2f" % time_in_service)+" horas"
    
    data = {'response_time': response_time_text, 'average_duration': average_duration_text, 'time_in_service': time_in_service_text}
    return HttpResponse(json.dumps(data))
